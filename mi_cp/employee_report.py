# import frappe
# import openpyxl
# from openpyxl.utils import get_column_letter
# from io import BytesIO

# def format_time_in_hours_and_minutes(minutes):
#     """Convert minutes into 'x hr, y min' format."""
#     hours = minutes // 60
#     mins = minutes % 60
#     if hours > 0:
#         return f"{hours} hr, {mins} min" if mins > 0 else f"{hours} hr"
#     else:
#         return f"{mins} min"

# @frappe.whitelist(allow_guest=True)
# def generate_employee_report(employee, start_date=None, end_date=None):
#     """
#     Generate and download a report for a specific employee with data in a single row.
#     """
#     if not frappe.db.exists("Employee", employee):
#         frappe.throw(f"Employee '{employee}' not found.")

#     # Set default date range if not provided
#     if not start_date:
#         start_date = "2024-12-01"
#     if not end_date:
#         end_date = "2024-12-31"

#     # Validate the date range
#     if frappe.utils.date_diff(end_date, start_date) < 0:
#         frappe.throw("End date cannot be before start date.")

#     num_days = frappe.utils.date_diff(end_date, start_date) + 1

#     # Fetch employee details
#     employee_data = frappe.db.get_value("Employee", employee, ["name", "employee_name", "department"], as_dict=True)

#     # Fetch shift details for the employee
#     shift_data = frappe.db.sql("""
#         SELECT DISTINCT 
#             shift, 
#             TIME(shift_start) AS shift_start, 
#             TIME(shift_end) AS shift_end
#         FROM `tabEmployee Checkin`
#         WHERE employee = %s
#         LIMIT 1
#     """, (employee,), as_dict=True)

#     # Initialize attendance data and metrics
#     total_present, total_absent, total_late, total_early = 0, 0, 0, 0
#     attendance_row = []

#     for i in range(num_days):
#         date = frappe.utils.add_days(start_date, i)

#         daily_data = frappe.db.sql("""
#             SELECT 
#                 MAX(CASE WHEN log_type = 'IN' THEN TIME(time) END) AS check_in,
#                 MAX(CASE WHEN log_type = 'OUT' THEN TIME(time) END) AS check_out,
#                 CASE
#                     WHEN MAX(CASE WHEN log_type = 'IN' THEN TIME(time) END) > TIME(shift_start) THEN 
#                         TIMESTAMPDIFF(MINUTE, TIME(shift_start), MAX(CASE WHEN log_type = 'IN' THEN TIME(time) END)) - 15
#                     ELSE 0
#                 END AS late_minutes,
#                 CASE
#                     WHEN MAX(CASE WHEN log_type = 'OUT' THEN TIME(time) END) < TIME(shift_end) THEN 
#                         TIMESTAMPDIFF(MINUTE, MAX(CASE WHEN log_type = 'OUT' THEN TIME(time) END), TIME(shift_end))
#                     ELSE 0
#                 END AS early_minutes
#             FROM `tabEmployee Checkin`
#             WHERE employee = %s AND DATE(time) = %s
#         """, (employee, date), as_dict=True)

#         if daily_data and daily_data[0]["check_in"]:
#             check_in = daily_data[0]["check_in"]
#             check_out = daily_data[0]["check_out"] or "--"
#             late_minutes = max(0, daily_data[0]["late_minutes"])
#             early_minutes = max(0, daily_data[0]["early_minutes"])

#             total_late += late_minutes
#             total_early += early_minutes

#             attendance_row.append(f"P ({check_in} - {check_out})")
#             total_present += 1
#         else:
#             attendance_row.append("A")
#             total_absent += 1

#     # Create Excel workbook
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = "Attendance Report"

#     # Define headers
#     headers = ["Emp Id", "Name", "Department", "Shift", "Shift Start", "Shift End"] + [
#         frappe.utils.formatdate(frappe.utils.add_days(start_date, i), "dd MMM") for i in range(num_days)
#     ] + ["Total Present", "Total Absent", "Total Late (hrs, mins)", "Total Early (hrs, mins)"]

#     # Write headers
#     for col_num, header in enumerate(headers, 1):
#         sheet.cell(row=1, column=col_num, value=header)

#     # Write data in a single row
#     row_num = 2
#     sheet.cell(row=row_num, column=1, value=employee_data["name"])
#     sheet.cell(row=row_num, column=2, value=employee_data["employee_name"])
#     sheet.cell(row=row_num, column=3, value=employee_data["department"])
#     sheet.cell(row=row_num, column=4, value=shift_data[0]["shift"] if shift_data else "Default Shift")
#     sheet.cell(row=row_num, column=5, value=shift_data[0]["shift_start"] if shift_data else "10:00:00")
#     sheet.cell(row=row_num, column=6, value=shift_data[0]["shift_end"] if shift_data else "19:00:00")

#     for i, attendance in enumerate(attendance_row, start=7):
#         sheet.cell(row=row_num, column=i, value=attendance)

#     # Summary metrics
#     sheet.cell(row=row_num, column=7 + num_days, value=total_present)
#     sheet.cell(row=row_num, column=8 + num_days, value=total_absent)
#     sheet.cell(row=row_num, column=9 + num_days, value=format_time_in_hours_and_minutes(total_late))
#     sheet.cell(row=row_num, column=10 + num_days, value=format_time_in_hours_and_minutes(total_early))

#     # Adjust column widths
#     for col_num in range(1, len(headers) + 1):
#         column_letter = get_column_letter(col_num)
#         sheet.column_dimensions[column_letter].width = 20

#     # Save workbook to a file-like object
#     output = BytesIO()
#     workbook.save(output)
#     output.seek(0)

#     # Serve the file for download
#     frappe.local.response.filecontent = output.getvalue()
#     frappe.local.response.type = "binary"
#     frappe.local.response.filename = f"{employee}_attendance_report.xlsx"



# id 3 
# import frappe
# import openpyxl
# from openpyxl.utils import get_column_letter
# from io import BytesIO

# def format_time_in_hours_and_minutes(minutes):
#     """Convert minutes into 'x hr, y min' format."""
#     hours = minutes // 60
#     mins = minutes % 60
#     if hours > 0:
#         return f"{hours} hr, {mins} min" if mins > 0 else f"{hours} hr"
#     else:
#         return f"{mins} min"

# @frappe.whitelist(allow_guest=True)
# def generate_employee_report(employee, month=None, year=None):
#     """
#     Generate and download a report for a specific employee with data in a single row.
#     """
#     if not frappe.db.exists("Employee", employee):
#         frappe.throw(f"Employee '{employee}' not found.")

#     # Validate and set default month and year if not provided
#     try:
#         if not month:
#             month = int(frappe.utils.nowdate().split("-")[1])  # Default to current month
#         else:
#             month = int(month)
#         if not year:
#             year = int(frappe.utils.nowdate().split("-")[0])  # Default to current year
#         else:
#             year = int(year)
#     except ValueError:
#         frappe.throw("Invalid month or year. Please provide numeric values.")

#     # Ensure month is in valid range
#     if not (1 <= month <= 12):
#         frappe.throw("Month must be between 1 and 12.")

#     # Calculate start_date and end_date for the given month and year
#     start_date = frappe.utils.get_first_day(frappe.utils.getdate(f"{year}-{month:02d}-01"))
#     end_date = frappe.utils.get_last_day(frappe.utils.getdate(f"{year}-{month:02d}-01"))

#     num_days = frappe.utils.date_diff(end_date, start_date) + 1

#     # Fetch employee details
#     employee_data = frappe.db.get_value("Employee", employee, ["name", "employee_name", "department"], as_dict=True)

#     # Fetch shift details for the employee
#     shift_data = frappe.db.sql(
#         """
#         SELECT DISTINCT 
#             shift, 
#             TIME(shift_start) AS shift_start, 
#             TIME(shift_end) AS shift_end
#         FROM `tabEmployee Checkin`
#         WHERE employee = %s
#         LIMIT 1
#         """, (employee,), as_dict=True)

#     # Initialize attendance data and metrics
#     total_present, total_absent, total_late, total_early = 0, 0, 0, 0
#     attendance_row = []

#     for i in range(num_days):
#         date = frappe.utils.add_days(start_date, i)

#         daily_data = frappe.db.sql(
#             """
#             SELECT 
#                 MAX(CASE WHEN log_type = 'IN' THEN TIME(time) END) AS check_in,
#                 MAX(CASE WHEN log_type = 'OUT' THEN TIME(time) END) AS check_out,
#                 CASE
#                     WHEN MAX(CASE WHEN log_type = 'IN' THEN TIME(time) END) > TIME(shift_start) THEN 
#                         TIMESTAMPDIFF(MINUTE, TIME(shift_start), MAX(CASE WHEN log_type = 'IN' THEN TIME(time) END)) - 15
#                     ELSE 0
#                 END AS late_minutes,
#                 CASE
#                     WHEN MAX(CASE WHEN log_type = 'OUT' THEN TIME(time) END) < TIME(shift_end) THEN 
#                         TIMESTAMPDIFF(MINUTE, MAX(CASE WHEN log_type = 'OUT' THEN TIME(time) END), TIME(shift_end))
#                     ELSE 0
#                 END AS early_minutes
#             FROM `tabEmployee Checkin`
#             WHERE employee = %s AND DATE(time) = %s
#             """, (employee, date), as_dict=True)

#         if daily_data and daily_data[0]["check_in"]:
#             # Mark the day as present
#             attendance_row.append("P")
#             total_present += 1
#             total_late += max(0, daily_data[0].get("late_minutes", 0))
#             total_early += max(0, daily_data[0].get("early_minutes", 0))
#         else:
#             # Mark the day as absent
#             attendance_row.append("A")
#             total_absent += 1

#     # Create Excel workbook
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = "Attendance Report"

#     # Define headers
#     headers = ["Emp Id", "Name", "Department", "Shift", "Shift Start", "Shift End"] + [
#         frappe.utils.formatdate(frappe.utils.add_days(start_date, i), "dd MMM") for i in range(num_days)
#     ] + ["Total Present", "Total Absent", "Total Late (hrs, mins)", "Total Early (hrs, mins)"]

#     # Write headers
#     for col_num, header in enumerate(headers, 1):
#         sheet.cell(row=1, column=col_num, value=header)

#     # Write data in a single row
#     row_num = 2
#     sheet.cell(row=row_num, column=1, value=employee_data["name"])
#     sheet.cell(row=row_num, column=2, value=employee_data["employee_name"])
#     sheet.cell(row=row_num, column=3, value=employee_data["department"])
#     sheet.cell(row=row_num, column=4, value=shift_data[0]["shift"] if shift_data else "Default Shift")
#     sheet.cell(row=row_num, column=5, value=shift_data[0]["shift_start"] if shift_data else "10:00:00")
#     sheet.cell(row=row_num, column=6, value=shift_data[0]["shift_end"] if shift_data else "19:00:00")

#     for i, attendance in enumerate(attendance_row, start=7):
#         sheet.cell(row=row_num, column=i, value=attendance)

#     # Summary metrics
#     sheet.cell(row=row_num, column=7 + num_days, value=total_present)
#     sheet.cell(row=row_num, column=8 + num_days, value=total_absent)
#     sheet.cell(row=row_num, column=9 + num_days, value=format_time_in_hours_and_minutes(total_late))
#     sheet.cell(row=row_num, column=10 + num_days, value=format_time_in_hours_and_minutes(total_early))

#     # Adjust column widths
#     for col_num in range(1, len(headers) + 1):
#         column_letter = get_column_letter(col_num)
#         sheet.column_dimensions[column_letter].width = 20

#     # Save workbook to a file-like object
#     output = BytesIO()
#     workbook.save(output)
#     output.seek(0)

#     # Serve the file for download
#     frappe.local.response.filecontent = output.getvalue()
#     frappe.local.response.type = "binary"
#     frappe.local.response.filename = f"{employee}_attendance_report_{month:02d}_{year}.xlsx"



import frappe
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from io import BytesIO

def format_time_in_hours_and_minutes(minutes):
    """Convert minutes into 'x hr, y min' format."""
    hours = minutes // 60
    mins = minutes % 60
    if hours > 0:
        return f"{hours} hr, {mins} min" if mins > 0 else f"{hours} hr"
    else:
        return f"{mins} min"

@frappe.whitelist()
def generate_employee_report(employee, month, year):
    """
    Generate and download a detailed attendance report for a specific employee.
    """
    # Validate and convert month and year to integers
    try:
        month = int(month)  # Convert month to integer
        year = int(year)    # Convert year to integer
    except ValueError:
        frappe.throw("Month and year must be numeric.")

    # Validate month and year range
    if not (1 <= month <= 12):
        frappe.throw("Month must be between 1 and 12.")
    if year < 1900 or year > 2100:
        frappe.throw("Year must be in a valid range.")

    # Ensure employee exists
    if not frappe.db.exists("Employee", employee):
        frappe.throw(f"Employee '{employee}' not found.")

    # Fetch employee details
    employee_data = frappe.db.get_value(
        "Employee", employee, ["name", "employee_name", "department"], as_dict=True
    )

    # Calculate date range for the given month and year
    start_date = frappe.utils.get_first_day(frappe.utils.getdate(f"{year}-{month:02d}-01"))
    end_date = frappe.utils.get_last_day(frappe.utils.getdate(f"{year}-{month:02d}-01"))
    num_days = frappe.utils.date_diff(end_date, start_date) + 1

    # Create workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Detailed Work Duration"

    # Add meta-information rows
    sheet.append(["Emp Id", employee_data["name"]])
    sheet.append(["Name", employee_data["employee_name"]])
    sheet.append(["Department", employee_data["department"]])
    sheet.append(["Location", "Default Location"])  # Replace with dynamic location if available
    sheet.append(["Shift", "General"])  # Replace "General" with dynamic shift data if available

    # Add header row for attendance
    sheet.append(["Days"] + [
        frappe.utils.formatdate(frappe.utils.add_days(start_date, i), "dd MMM")
        for i in range(num_days)
    ])

    # Initialize rows for attendance details
    attendance_row = ["Attendance"]
    intime_row = ["InTime"]
    outtime_row = ["OutTime"]
    duration_row = ["Duration"]
    late_by_row = ["Late By"]
    early_by_row = ["Early By"]

    # Populate attendance data
    for i in range(num_days):
        date = frappe.utils.add_days(start_date, i)
        daily_data = frappe.db.sql(
            """
            SELECT 
                MAX(CASE WHEN log_type = 'IN' THEN TIME(time) END) AS check_in,
                MAX(CASE WHEN log_type = 'OUT' THEN TIME(time) END) AS check_out,
                CASE
                    WHEN MAX(CASE WHEN log_type = 'IN' THEN TIME(time) END) > TIME('09:00:00') THEN 
                        TIMESTAMPDIFF(MINUTE, TIME('09:00:00'), MAX(CASE WHEN log_type = 'IN' THEN TIME(time) END))
                    ELSE 0
                END AS late_minutes,
                CASE
                    WHEN MAX(CASE WHEN log_type = 'OUT' THEN TIME(time) END) < TIME('18:00:00') THEN 
                        TIMESTAMPDIFF(MINUTE, MAX(CASE WHEN log_type = 'OUT' THEN TIME(time) END), TIME('18:00:00'))
                    ELSE 0
                END AS early_minutes
            FROM `tabEmployee Checkin`
            WHERE employee = %s AND DATE(time) = %s
            """,
            (employee, date), as_dict=True
        )

        if daily_data and daily_data[0]["check_in"]:
            attendance_row.append("P")
            intime_row.append(daily_data[0]["check_in"] or "-")
            outtime_row.append(daily_data[0]["check_out"] or "-")
            duration = "-"
            if daily_data[0]["check_in"] and daily_data[0]["check_out"]:
                duration = str(frappe.utils.time_diff_in_seconds(daily_data[0]["check_out"], daily_data[0]["check_in"]) // 60) + " mins"
            duration_row.append(duration)
            late_by_row.append(format_time_in_hours_and_minutes(daily_data[0].get("late_minutes", 0)))
            early_by_row.append(format_time_in_hours_and_minutes(daily_data[0].get("early_minutes", 0)))
        else:
            attendance_row.append("A")
            intime_row.append("-")
            outtime_row.append("-")
            duration_row.append("-")
            late_by_row.append("-")
            early_by_row.append("-")

    # Append all rows to the sheet
    sheet.append(attendance_row)
    sheet.append(intime_row)
    sheet.append(outtime_row)
    sheet.append(duration_row)
    sheet.append(late_by_row)
    sheet.append(early_by_row)

    # Adjust column widths
    for col_num in range(1, len(attendance_row) + 1):
        sheet.column_dimensions[get_column_letter(col_num)].width = 15

    # Save workbook to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Return as downloadable file
    frappe.local.response.filecontent = output.getvalue()
    frappe.local.response.type = "binary"
    frappe.local.response.filename = f"{employee}_attendance_{month:02d}_{year}.xlsx"
