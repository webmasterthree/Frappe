import frappe
from calendar import monthrange
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter

def format_time_in_hours_and_minutes(minutes):
    """Convert minutes into 'x hr, y min' format."""
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours} hr, {mins} min" if hours > 0 else f"{mins} min"

@frappe.whitelist()
def generate_employee_excel_report(month, year, employee):
    """
    Generate and download the detailed attendance report in Excel format.
    """
    # Validate inputs
    if not employee or not month or not year:
        frappe.throw("Please provide Employee, Month, and Year to generate the report.")

    # Calculate start_date and end_date
    start_date = f"{year}-{month.zfill(2)}-01"
    days_in_month = monthrange(int(year), int(month))[1]
    end_date = f"{year}-{month.zfill(2)}-{days_in_month}"

    # Fetch employee details
    employee_data = frappe.db.get_value("Employee", employee, ["name", "employee_name", "department"], as_dict=True)
    if not employee_data:
        frappe.throw("Employee details not found!")

    # Fetch shift details
    shift_data = frappe.db.sql("""
        SELECT DISTINCT shift, TIME(shift_start) AS shift_start, TIME(shift_end) AS shift_end
        FROM `tabEmployee Checkin`
        WHERE employee = %s
        LIMIT 1
    """, (employee,), as_dict=True)
    shift_info = shift_data[0] if shift_data else {"shift": "Default", "shift_start": None, "shift_end": None}

    # Generate daily data
    num_days = frappe.utils.date_diff(end_date, start_date) + 1
    report_data = []
    for i in range(num_days):
        date = frappe.utils.add_days(start_date, i)
        daily_data = frappe.db.sql("""
            SELECT 
                MAX(CASE WHEN log_type = 'IN' THEN TIME(time) END) AS check_in,
                MAX(CASE WHEN log_type = 'OUT' THEN TIME(time) END) AS check_out,
                CASE
                    WHEN MAX(CASE WHEN log_type = 'IN' THEN TIME(time) END) IS NOT NULL THEN 'P'
                    ELSE 'A'
                END AS status
            FROM `tabEmployee Checkin`
            WHERE employee = %s AND DATE(time) = %s
        """, (employee, date), as_dict=True)

        check_in = str(daily_data[0]["check_in"]) if daily_data and daily_data[0]["check_in"] else None
        check_out = str(daily_data[0]["check_out"]) if daily_data and daily_data[0]["check_out"] else None
        status = daily_data[0]["status"] if daily_data else "A"

        work_duration_minutes = (
            calculate_duration_in_minutes(check_in, check_out)
            if check_in and check_out else 0
        )
        work_duration = format_time_in_hours_and_minutes(work_duration_minutes)

        report_data.append({
            "date": date,
            "status": status,
            "check_in": check_in,
            "check_out": check_out,
            "work_duration": work_duration,
        })

    # Generate Excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance Report"

    # Define headers
    headers = ["Date", "Status", "Check-in Time", "Check-out Time", "Work Duration"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Add data rows
    for row_num, row_data in enumerate(report_data, start=2):
        sheet.cell(row=row_num, column=1, value=row_data["date"])
        sheet.cell(row=row_num, column=2, value=row_data["status"])
        sheet.cell(row=row_num, column=3, value=row_data["check_in"])
        sheet.cell(row=row_num, column=4, value=row_data["check_out"])
        sheet.cell(row=row_num, column=5, value=row_data["work_duration"])

    # Adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 20

    # Save to BytesIO and return
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    frappe.local.response.filecontent = output.getvalue()
    frappe.local.response.type = "binary"
    frappe.local.response.filename = f"{employee}_attendance_report_{month}_{year}.xlsx"

