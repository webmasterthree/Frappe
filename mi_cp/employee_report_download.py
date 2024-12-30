import frappe
from frappe.utils.response import download
import openpyxl
from openpyxl.utils import get_column_letter

@frappe.whitelist()
def employee_report_download(employee):
    """
    Generate and download an Excel file for a specific employee's report.
    """
    # Validate employee
    if not frappe.db.exists("Employee", employee):
        frappe.throw(f"Employee '{employee}' not found.")

    # Fetch attendance data for the employee
    attendance_data = frappe.db.sql("""
        SELECT 
            attendance_date, status, check_in_time, check_out_time
        FROM `tabAttendance`
        WHERE employee = %s
        ORDER BY attendance_date
    """, (employee,), as_dict=True)

    if not attendance_data:
        frappe.throw(f"No attendance data found for employee '{employee}'.")

    # Create Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance Report"

    # Set column headers
    headers = ["Date", "Status", "Check-in Time", "Check-out Time"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Populate sheet with data
    for row_num, record in enumerate(attendance_data, start=2):
        sheet.cell(row=row_num, column=1, value=record["attendance_date"])
        sheet.cell(row=row_num, column=2, value=record["status"])
        sheet.cell(row=row_num, column=3, value=record.get("check_in_time", "N/A"))
        sheet.cell(row=row_num, column=4, value=record.get("check_out_time", "N/A"))

    # Adjust column widths
    for col_num, _ in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 20

    # Save workbook to a file-like object
    from io import BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Provide the file as a downloadable response
    filename = f"{employee}_attendance_report.xlsx"
    return download(output, filename)
